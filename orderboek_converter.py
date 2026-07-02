from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


EXPORT_COLUMNS = [
    "Relatie",
    "Relatie tel prive",
    "Relatie mobiel",
    "Relatie email",
    "Relatie postcode",
    "Relatie straat",
    "Relatie huisnr.",
    "Relatie huisnr. toev.",
    "Relatie plaats",
    "Merk",
    "Autoomschrijving",
    "Afleveringmodel",
    "Kenteken",
    "Chassisnummer",
    "Relatie geslacht",
    "Chassisnummer controle",
]

SOURCE_COLUMNS = [c for c in EXPORT_COLUMNS if c != "Chassisnummer controle"]

ALIASES = {
    "Relatie tel prive": ["Relatie tel privé", "Relatie telefoon prive", "Relatie telefoon privé"],
    "Relatie mobiel": ["Mobiel", "Relatie mobiele telefoon"],
    "Relatie email": ["Relatie e-mail", "Email", "E-mail"],
    "Relatie postcode": ["Postcode"],
    "Relatie straat": ["Straat"],
    "Relatie huisnr.": ["Relatie huisnummer", "Huisnummer", "Huisnr."],
    "Relatie huisnr. toev.": ["Relatie huisnummer toev.", "Relatie huisnr toev", "Toevoeging"],
    "Relatie plaats": ["Plaats"],
    "Autoomschrijving": ["Auto omschrijving", "Omschrijving auto", "Automodel"],
    "Afleveringmodel": ["Aflevering model", "Aflevermodel"],
    "Relatie geslacht": ["Geslacht"],
}


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = str(value).strip()
    if s.endswith(".0"):
        try:
            number = float(s)
            if number.is_integer():
                return str(int(number))
        except Exception:
            pass
    return s


def _normalize_header(value: Any) -> str:
    s = _safe_str(value).lower()
    s = s.replace("é", "e").replace("è", "e")
    return re.sub(r"[^a-z0-9]+", "", s)


def _source_names_for(column: str) -> list[str]:
    return [column, *ALIASES.get(column, [])]


def _find_header_row(raw: pd.DataFrame) -> int:
    wanted = {_normalize_header(name) for column in SOURCE_COLUMNS for name in _source_names_for(column)}
    best_idx = 0
    best_score = -1
    max_scan = min(len(raw), 20)

    for idx in range(max_scan):
        values = [_normalize_header(v) for v in raw.iloc[idx].tolist()]
        score = sum(1 for v in values if v in wanted)
        if score > best_score:
            best_idx = idx
            best_score = score

    return best_idx


def _read_orderboek(file_path: str | Path) -> tuple[pd.DataFrame, dict[str, str]]:
    raw = pd.read_excel(file_path, header=None, dtype=object)
    if raw.empty:
        return pd.DataFrame(), {}

    header_idx = _find_header_row(raw)
    headers = [_safe_str(v) for v in raw.iloc[header_idx].tolist()]
    data = raw.iloc[header_idx + 1 :].copy()
    data.columns = headers
    data = data.dropna(how="all")

    source_map: dict[str, str] = {}
    normalized_headers = {_normalize_header(header): header for header in headers if _safe_str(header)}

    for column in SOURCE_COLUMNS:
        for source_name in _source_names_for(column):
            found = normalized_headers.get(_normalize_header(source_name))
            if found is not None:
                source_map[column] = found
                break

    return data, source_map


def clean_phone(value: Any) -> str:
    s = _safe_str(value)
    if not s:
        return ""

    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    s = re.sub(r"[\s\-\(\)]", "", s)
    if s.startswith("+31"):
        s = "0" + s[3:]
    elif s.startswith("0031"):
        s = "0" + s[4:]

    digits = re.sub(r"\D", "", s)
    if len(digits) == 9 and not digits.startswith("0"):
        digits = "0" + digits
    return digits


def clean_postcode(value: Any) -> str:
    s = re.sub(r"\s+", "", _safe_str(value).upper())
    match = re.fullmatch(r"(\d{4})([A-Z]{2})", s)
    if match:
        return f"{match.group(1)} {match.group(2)}"
    return _safe_str(value).upper()


def clean_kenteken(value: Any) -> str:
    return re.sub(r"\s+", "", _safe_str(value).upper())


def clean_chassisnummer(value: Any) -> str:
    return _safe_str(value).upper()


def chassisnummer_controle(value: Any) -> str:
    chassis = clean_chassisnummer(value)
    if not chassis:
        return "Ontbreekt"
    if len(chassis) == 17:
        return "OK"
    if len(chassis) < 17:
        return "Te kort"
    return "Te lang"


def _build_export_frame(data: pd.DataFrame, source_map: dict[str, str]) -> pd.DataFrame:
    rows: dict[str, list[str]] = {}
    for column in SOURCE_COLUMNS:
        source = source_map.get(column)
        if source is None:
            rows[column] = [""] * len(data)
        else:
            rows[column] = [_safe_str(value) for value in data[source].tolist()]

    out = pd.DataFrame(rows, columns=SOURCE_COLUMNS)
    out["Relatie tel prive"] = out["Relatie tel prive"].map(clean_phone)
    out["Relatie mobiel"] = out["Relatie mobiel"].map(clean_phone)
    out["Relatie postcode"] = out["Relatie postcode"].map(clean_postcode)
    out["Kenteken"] = out["Kenteken"].map(clean_kenteken)
    out["Chassisnummer"] = out["Chassisnummer"].map(clean_chassisnummer)
    out["Chassisnummer controle"] = out["Chassisnummer"].map(chassisnummer_controle)
    return out[EXPORT_COLUMNS]


def _remove_duplicate_customers(out: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    seen: set[tuple[str, str, str]] = set()
    keep_indexes: list[int] = []
    duplicates = 0

    for idx, row in out.iterrows():
        key = (
            _safe_str(row.get("Relatie email")).lower(),
            clean_phone(row.get("Relatie mobiel")),
            clean_chassisnummer(row.get("Chassisnummer")),
        )
        if not any(key):
            keep_indexes.append(idx)
            continue
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)
        keep_indexes.append(idx)

    return out.loc[keep_indexes].reset_index(drop=True), duplicates


def _write_orderboek_excel(out: pd.DataFrame, output_path: Path) -> None:
    out.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    text_columns = {"Relatie tel prive", "Relatie mobiel"}
    header_indexes = {
        str(cell.value or ""): cell.column
        for cell in ws[1]
    }

    for header in text_columns:
        column_idx = header_indexes.get(header)
        if not column_idx:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=column_idx)
            cell.number_format = "@"
            cell.value = "" if cell.value is None else str(cell.value)

    wb.save(output_path)


def convert_orderboek(
    file_path: str | Path,
    output_path: str | Path,
    remove_duplicates: bool = False,
) -> dict[str, int | str]:
    data, source_map = _read_orderboek(file_path)
    out = _build_export_frame(data, source_map)

    imported_rows = len(out)
    duplicates_removed = 0
    if remove_duplicates:
        out, duplicates_removed = _remove_duplicate_customers(out)

    missing_columns = [column for column in SOURCE_COLUMNS if column not in source_map]
    missing_chassis = int((out["Chassisnummer"].astype(str).str.strip() == "").sum())

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_orderboek_excel(out, output_path)

    return {
        "imported_rows": imported_rows,
        "exported_rows": len(out),
        "missing_columns_count": len(missing_columns),
        "duplicates_removed": duplicates_removed,
        "missing_chassis_count": missing_chassis,
        "missing_columns": ", ".join(missing_columns),
    }
